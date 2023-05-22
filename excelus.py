from openpyxl import load_workbook


def r_excel():
    fn = 'attachments/pars_result.xlsx'
    wb = load_workbook(fn, data_only=True)

    ids_all = []
    sheets = ['mvideo', 'eldorado', 'vse_instrumenti', 'ozon', 'wb']

    for sht in sheets:
        sheet = wb[sht]
        # max_rows = sheet.max_row
        max_rows = 6
        ids = []
        for i in range(2, max_rows + 1):
            value = sheet.cell(row=i, column=3).value
            ids.append(value)
        ids_all.append(ids)

    wb.close()
    return ids_all


def w_excel(res_all):
    print(f'[!] Начало заполнение файла')
    fn = 'attachments/pars_result.xlsx'
    wb = load_workbook(fn, data_only=True)

    sheets = ['mvideo', 'vse_instrumenti', 'ozon']
    j = 0

    for sht in sheets:

        sheet = wb[sht]
        # max_rows = sheet.max_row
        max_rows = 6

        res_ava = res_all[j][0]
        res_cou = res_all[j][1]

        for i in range(2, max_rows + 1):
            col_avaible = 'D' + str(i)
            col_count = 'E' + str(i)
            sheet[col_avaible] = res_ava[i - 2]
            sheet[col_count] = res_cou[i - 2]

        j = j + 1

    shet = wb['all']
    # max_rows = shet.max_row
    max_rows = 6

    cou_mv = res_all[0][1]
    cou_vi = res_all[1][1]
    cou_oz = res_all[2][1]

    for r in range(2, max_rows + 1):
        col_m = 'C' + str(r)
        col_v = 'E' + str(r)
        col_o = 'F' + str(r)
        shet[col_m] = cou_mv[r - 2]
        shet[col_v] = cou_vi[r - 2]
        shet[col_o] = cou_oz[r - 2]

    wb.save(fn)
    wb.close()
    print(f'[!] Запись окончена')
